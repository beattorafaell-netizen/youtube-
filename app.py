#!/usr/bin/env python3
"""
YouTube Upload Scheduler
=========================
- Interface web em http://localhost:5000
- Bot no Telegram para agendar sem abrir o site
- Exportar histórico em Excel
- Edição de título/descrição após agendar
- Thumbnail personalizada
- Uploads recorrentes
- Notificação no Telegram após cada upload

Dependências:
    pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib flask requests openpyxl

Rodar:
    python app.py
"""

import os
import io
import time
import json
import pickle
import shutil
import logging
import threading
import requests
from pathlib import Path
from datetime import datetime, timedelta

from flask import Flask, request, jsonify, send_from_directory, send_file
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
#  CONFIGURAÇÕES
# ─────────────────────────────────────────────

PASTA_UPLOADS    = Path("videos")
PASTA_ENVIADOS   = Path("enviados")
PASTA_THUMBS     = Path("thumbs")
CREDENTIALS_FILE = "client_secrets.json"
TOKEN_FILE       = "token.pickle"
HISTORICO_FILE   = "historico.json"

TELEGRAM_TOKEN   = "8608511631:AAFG9qPvifnJziLCL4UE-rMwMZWH37d27pk"
TELEGRAM_CHAT_ID = "8144548560"

SCOPES = [
    "https://www.googleapis.com/auth/youtube.upload",
    "https://www.googleapis.com/auth/youtube",
]

fila_agendados = []
fila_lock = threading.Lock()

# Estado do bot Telegram (para conversa multi-etapa)
tg_state = {}
tg_state_lock = threading.Lock()

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__, static_folder=".")

# ─────────────────────────────────────────────
#  HISTÓRICO
# ─────────────────────────────────────────────

def carregar_historico() -> list:
    if Path(HISTORICO_FILE).exists():
        with open(HISTORICO_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def salvar_historico(historico: list):
    with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
        json.dump(historico, f, ensure_ascii=False, indent=2)

def adicionar_historico(titulo: str, url: str, privacidade: str):
    historico = carregar_historico()
    historico.insert(0, {
        "titulo":      titulo,
        "url":         url,
        "privacidade": privacidade,
        "data":        datetime.now().isoformat(),
    })
    salvar_historico(historico)

# ─────────────────────────────────────────────
#  TELEGRAM — ENVIO
# ─────────────────────────────────────────────

def tg_send(chat_id: str, texto: str, reply_markup=None):
    try:
        payload = {"chat_id": chat_id, "text": texto, "parse_mode": "HTML"}
        if reply_markup:
            payload["reply_markup"] = json.dumps(reply_markup)
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json=payload, timeout=10
        )
    except Exception as e:
        log.warning("Telegram send error: %s", e)

# ─────────────────────────────────────────────
#  TELEGRAM — BOT (polling)
# ─────────────────────────────────────────────

AJUDA = (
    "📋 <b>Comandos disponíveis:</b>\n\n"
    "/agendar — Agendar um novo upload\n"
    "/fila — Ver vídeos na fila\n"
    "/historico — Ver últimos uploads\n"
    "/ajuda — Ver esta mensagem"
)

def tg_handle(update: dict):
    msg = update.get("message") or update.get("edited_message")
    if not msg:
        return
    chat_id = str(msg["chat"]["id"])
    text = msg.get("text", "").strip()

    with tg_state_lock:
        estado = tg_state.get(chat_id, {})

    # Comandos principais
    if text == "/start" or text == "/ajuda":
        tg_send(chat_id, f"👋 Olá! Bem-vindo ao YouTube Scheduler Bot!\n\n{AJUDA}")
        return

    if text == "/fila":
        with fila_lock:
            pendentes = [i for i in fila_agendados if i["status"] in ("aguardando", "enviando")]
        if not pendentes:
            tg_send(chat_id, "📭 Nenhum vídeo na fila no momento.")
        else:
            linhas = [f"📋 <b>Fila ({len(pendentes)} vídeo(s)):</b>\n"]
            for i, item in enumerate(pendentes, 1):
                h = datetime.fromisoformat(item["horario"]).strftime("%d/%m %H:%M")
                linhas.append(f"{i}. <b>{item['titulo']}</b>\n   📅 {h} · {item['status']}")
            tg_send(chat_id, "\n".join(linhas))
        return

    if text == "/historico":
        hist = carregar_historico()[:5]
        if not hist:
            tg_send(chat_id, "📭 Nenhum vídeo postado ainda.")
        else:
            linhas = ["📋 <b>Últimos uploads:</b>\n"]
            for item in hist:
                d = datetime.fromisoformat(item["data"]).strftime("%d/%m %H:%M")
                linhas.append(f"🎬 <b>{item['titulo']}</b>\n   📅 {d}\n   🔗 {item['url']}")
            tg_send(chat_id, "\n".join(linhas))
        return

    if text == "/agendar":
        with tg_state_lock:
            tg_state[chat_id] = {"etapa": "aguardando_titulo"}
        tg_send(chat_id,
            "🎬 <b>Novo agendamento</b>\n\n"
            "Vamos configurar seu upload em poucos passos.\n\n"
            "<b>Etapa 1/4 — Título</b>\n"
            "Digite o título do vídeo:"
        )
        return

    # Fluxo de agendamento por etapas
    with tg_state_lock:
        estado = tg_state.get(chat_id, {})

    if estado.get("etapa") == "aguardando_titulo":
        with tg_state_lock:
            tg_state[chat_id] = {"etapa": "aguardando_descricao", "titulo": text}
        tg_send(chat_id,
            f"✅ Título: <b>{text}</b>\n\n"
            "<b>Etapa 2/4 — Descrição</b>\n"
            "Digite a descrição (ou envie <code>-</code> para pular):"
        )
        return

    if estado.get("etapa") == "aguardando_descricao":
        descricao = "" if text == "-" else text
        with tg_state_lock:
            tg_state[chat_id] = {**estado, "etapa": "aguardando_horario", "descricao": descricao}
        tg_send(chat_id,
            "<b>Etapa 3/4 — Horário</b>\n"
            "Digite a data e hora do upload no formato:\n"
            "<code>DD/MM/AAAA HH:MM</code>\n\n"
            "Exemplo: <code>25/12/2025 18:00</code>\n"
            "Ou envie <code>agora</code> para postar imediatamente:"
        )
        return

    if estado.get("etapa") == "aguardando_horario":
        try:
            if text.lower() == "agora":
                horario = datetime.now()
            else:
                horario = datetime.strptime(text, "%d/%m/%Y %H:%M")
        except ValueError:
            tg_send(chat_id, "❌ Formato inválido! Use: <code>DD/MM/AAAA HH:MM</code>")
            return

        with tg_state_lock:
            tg_state[chat_id] = {**estado, "etapa": "aguardando_privacidade", "horario": horario.isoformat()}

        tg_send(chat_id,
            f"✅ Horário: <b>{horario.strftime('%d/%m/%Y às %H:%M')}</b>\n\n"
            "<b>Etapa 4/4 — Privacidade</b>\n"
            "Escolha enviando o número:\n\n"
            "1️⃣ Público\n2️⃣ Não listado\n3️⃣ Privado"
        )
        return

    if estado.get("etapa") == "aguardando_privacidade":
        mapa = {"1": "public", "2": "unlisted", "3": "private"}
        priv = mapa.get(text)
        if not priv:
            tg_send(chat_id, "❌ Envie 1, 2 ou 3.")
            return

        with tg_state_lock:
            tg_state[chat_id] = {**estado, "etapa": "aguardando_video", "privacidade": priv}

        priv_label = {"public": "🌍 Público", "unlisted": "🔗 Não listado", "private": "🔒 Privado"}[priv]
        tg_send(chat_id,
            f"✅ Privacidade: <b>{priv_label}</b>\n\n"
            "📁 <b>Última etapa!</b>\n"
            "Agora coloque o arquivo de vídeo na pasta <code>videos/</code> "
            "e envie o nome do arquivo (com extensão).\n\n"
            "Exemplo: <code>meu_video.mp4</code>"
        )
        return

    if estado.get("etapa") == "aguardando_video":
        caminho = PASTA_UPLOADS / text
        if not caminho.exists():
            tg_send(chat_id,
                f"❌ Arquivo <code>{text}</code> não encontrado em <code>videos/</code>.\n"
                "Verifique o nome e tente novamente."
            )
            return

        item = {
            "id":            int(time.time() * 1000),
            "titulo":        estado["titulo"],
            "descricao":     estado.get("descricao", ""),
            "horario":       estado["horario"],
            "privacidade":   estado["privacidade"],
            "recorrente":    False,
            "recorrencia":   "diario",
            "caminho":       str(caminho),
            "arquivo":       text,
            "thumb_caminho": None,
            "status":        "aguardando",
            "url":           None,
            "url_historico": [],
            "ultimo_envio":  None,
            "erro":          None,
        }

        with fila_lock:
            fila_agendados.append(item)

        with tg_state_lock:
            tg_state.pop(chat_id, None)

        h = datetime.fromisoformat(estado["horario"]).strftime("%d/%m/%Y às %H:%M")
        tg_send(chat_id,
            f"✅ <b>Agendado com sucesso!</b>\n\n"
            f"🎬 <b>{estado['titulo']}</b>\n"
            f"📅 {h}\n"
            f"📁 {text}\n\n"
            "Você receberá uma notificação quando o upload for concluído!"
        )
        return

    # Mensagem não reconhecida
    tg_send(chat_id, f"❓ Comando não reconhecido.\n\n{AJUDA}")

def tg_polling():
    """Polling do Telegram para receber mensagens."""
    log.info("Bot Telegram iniciado.")
    offset = None
    while True:
        try:
            params = {"timeout": 30, "allowed_updates": ["message"]}
            if offset:
                params["offset"] = offset
            resp = requests.get(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates",
                params=params, timeout=35
            )
            data = resp.json()
            if data.get("ok"):
                for update in data.get("result", []):
                    offset = update["update_id"] + 1
                    try:
                        tg_handle(update)
                    except Exception as e:
                        log.error("Erro ao processar update Telegram: %s", e)
        except Exception as e:
            log.warning("Erro no polling Telegram: %s", e)
            time.sleep(5)

# ─────────────────────────────────────────────
#  AUTENTICAÇÃO YOUTUBE
# ─────────────────────────────────────────────

def autenticar():
    credenciais = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "rb") as f:
            credenciais = pickle.load(f)
    if not credenciais or not credenciais.valid:
        if credenciais and credenciais.expired and credenciais.refresh_token:
            credenciais.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            credenciais = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "wb") as f:
            pickle.dump(credenciais, f)
    return build("youtube", "v3", credentials=credenciais)

# ─────────────────────────────────────────────
#  UPLOAD
# ─────────────────────────────────────────────

def fazer_upload(youtube, item: dict) -> dict:
    caminho_video = Path(item["caminho"])
    if not caminho_video.exists():
        return {"ok": False, "erro": f"Arquivo não encontrado: {caminho_video}"}

    corpo = {
        "snippet": {"title": item["titulo"], "description": item["descricao"], "categoryId": "22"},
        "status":  {"privacyStatus": item["privacidade"]},
    }
    media = MediaFileUpload(str(caminho_video), mimetype="video/*", resumable=True, chunksize=10*1024*1024)

    try:
        req = youtube.videos().insert(part="snippet,status", body=corpo, media_body=media)
        resposta = None
        while resposta is None:
            _, resposta = req.next_chunk()

        video_id = resposta.get("id", "")
        url = f"https://youtu.be/{video_id}"

        # Thumbnail
        thumb = item.get("thumb_caminho")
        if thumb and Path(thumb).exists():
            try:
                youtube.thumbnails().set(videoId=video_id, media_body=MediaFileUpload(thumb, mimetype="image/*")).execute()
            except HttpError as e:
                log.warning("Thumbnail falhou: %s", e)

        # Move vídeo
        if not item.get("recorrente"):
            destino = PASTA_ENVIADOS / caminho_video.name
            if destino.exists():
                destino = PASTA_ENVIADOS / f"{caminho_video.stem}__{int(time.time())}{caminho_video.suffix}"
            shutil.move(str(caminho_video), str(destino))

        adicionar_historico(item["titulo"], url, item["privacidade"])

        priv_emoji = {"public": "🌍", "unlisted": "🔗", "private": "🔒"}.get(item["privacidade"], "")
        tg_send(TELEGRAM_CHAT_ID,
            f"✅ <b>Vídeo postado no YouTube!</b>\n\n"
            f"🎬 <b>{item['titulo']}</b>\n"
            f"{priv_emoji} {item['privacidade'].capitalize()}\n"
            f"🔗 {url}"
        )
        return {"ok": True, "video_id": video_id, "url": url}

    except HttpError as e:
        tg_send(TELEGRAM_CHAT_ID, f"❌ <b>Falha no upload!</b>\n🎬 {item['titulo']}\n⚠️ {e}")
        return {"ok": False, "erro": str(e)}
    except Exception as e:
        return {"ok": False, "erro": str(e)}

# ─────────────────────────────────────────────
#  PRÓXIMO HORÁRIO RECORRENTE
# ─────────────────────────────────────────────

def proximo_horario(item: dict) -> str:
    base = datetime.fromisoformat(item["horario"])
    agora = datetime.now()
    delta = timedelta(weeks=1) if item.get("recorrencia") == "semanal" else timedelta(days=1)
    while base <= agora:
        base += delta
    return base.isoformat()

# ─────────────────────────────────────────────
#  WORKER
# ─────────────────────────────────────────────

def worker():
    log.info("Worker iniciado.")
    youtube = autenticar()
    while True:
        agora = datetime.now()
        with fila_lock:
            for item in fila_agendados:
                if item["status"] != "aguardando":
                    continue
                if agora < datetime.fromisoformat(item["horario"]):
                    continue
                item["status"] = "enviando"
                resultado = fazer_upload(youtube, item)
                if resultado["ok"]:
                    item["url"] = resultado["url"]
                    item["ultimo_envio"] = datetime.now().isoformat()
                    if item.get("recorrente"):
                        item["url_historico"].append(resultado["url"])
                        item["horario"] = proximo_horario(item)
                        item["status"] = "aguardando"
                    else:
                        item["status"] = "enviado"
                else:
                    item["status"] = "erro"
                    item["erro"] = resultado["erro"]
        time.sleep(30)

# ─────────────────────────────────────────────
#  ROTAS FLASK
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/agendar", methods=["POST"])
def agendar():
    if "video" not in request.files:
        return jsonify({"ok": False, "erro": "Nenhum vídeo enviado"}), 400

    arquivo     = request.files["video"]
    titulo      = request.form.get("titulo", arquivo.filename.rsplit(".", 1)[0])
    descricao   = request.form.get("descricao", "")
    horario     = request.form.get("horario", datetime.now().isoformat())
    privacidade = request.form.get("privacidade", "public")
    recorrente  = request.form.get("recorrente", "false") == "true"
    recorrencia = request.form.get("recorrencia", "diario")

    for p in [PASTA_UPLOADS, PASTA_ENVIADOS, PASTA_THUMBS]:
        p.mkdir(exist_ok=True)

    caminho = PASTA_UPLOADS / arquivo.filename
    arquivo.save(str(caminho))

    thumb_caminho = None
    if "thumbnail" in request.files:
        thumb = request.files["thumbnail"]
        if thumb.filename:
            thumb_caminho = str(PASTA_THUMBS / f"{int(time.time())}_{thumb.filename}")
            thumb.save(thumb_caminho)

    item = {
        "id": int(time.time() * 1000), "titulo": titulo, "descricao": descricao,
        "horario": horario, "privacidade": privacidade, "recorrente": recorrente,
        "recorrencia": recorrencia, "caminho": str(caminho), "arquivo": arquivo.filename,
        "thumb_caminho": thumb_caminho, "status": "aguardando",
        "url": None, "url_historico": [], "ultimo_envio": None, "erro": None,
    }

    with fila_lock:
        fila_agendados.append(item)
    return jsonify({"ok": True, "item": item})

@app.route("/editar/<int:item_id>", methods=["PUT"])
def editar(item_id):
    data = request.json or {}
    with fila_lock:
        for item in fila_agendados:
            if item["id"] == item_id and item["status"] == "aguardando":
                if "titulo" in data:
                    item["titulo"] = data["titulo"]
                if "descricao" in data:
                    item["descricao"] = data["descricao"]
                if "horario" in data:
                    item["horario"] = data["horario"]
                if "privacidade" in data:
                    item["privacidade"] = data["privacidade"]
                return jsonify({"ok": True, "item": item})
    return jsonify({"ok": False, "erro": "Item não encontrado ou não editável"}), 404

@app.route("/fila")
def listar_fila():
    with fila_lock:
        return jsonify(fila_agendados)

@app.route("/historico")
def listar_historico():
    return jsonify(carregar_historico())

@app.route("/exportar-excel")
def exportar_excel():
    historico = carregar_historico()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico de Uploads"

    # Cabeçalho
    headers = ["Título", "Data/Hora", "Privacidade", "URL"]
    header_fill = PatternFill("solid", fgColor="FF2D2D")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row, item in enumerate(historico, 2):
        data_fmt = datetime.fromisoformat(item["data"]).strftime("%d/%m/%Y %H:%M")
        priv_map = {"public": "Público", "unlisted": "Não listado", "private": "Privado"}
        ws.cell(row=row, column=1, value=item["titulo"])
        ws.cell(row=row, column=2, value=data_fmt)
        ws.cell(row=row, column=3, value=priv_map.get(item["privacidade"], item["privacidade"]))
        ws.cell(row=row, column=4, value=item["url"])

    # Larguras
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 35

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"historico_youtube_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

@app.route("/cancelar/<int:item_id>", methods=["DELETE"])
def cancelar(item_id):
    with fila_lock:
        for item in fila_agendados:
            if item["id"] == item_id and item["status"] == "aguardando":
                item["status"] = "cancelado"
                Path(item["caminho"]).unlink(missing_ok=True)
                return jsonify({"ok": True})
    return jsonify({"ok": False, "erro": "Item não encontrado"}), 404

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    for p in [PASTA_UPLOADS, PASTA_ENVIADOS, PASTA_THUMBS]:
        p.mkdir(exist_ok=True)

    threading.Thread(target=worker, daemon=True).start()
    threading.Thread(target=tg_polling, daemon=True).start()

    log.info("=" * 50)
    log.info("  Acesse: http://localhost:5000")
    log.info("  Bot Telegram: @meuytupload_bot")
    log.info("=" * 50)
    app.run(debug=False, port=5000)
